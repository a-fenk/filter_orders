import datetime

from openpyxl import load_workbook, Workbook

from config import Config
from utils import count_unique_words_with_registry, count_duplicates, create_registry, count_unique_words_in_compare, \
    count_duplicates_between_rows, tokenize, stemmer


class Reverser:
    def __init__(self, obj):
        self.obj = obj

    def __eq__(self, other):
        return other.obj == self.obj

    def __lt__(self, other):
        return other.obj < self.obj


class FilteredRow:
    def __init__(self, value: str, main_row, registry: dict):
        self.value = value
        self.main_row = main_row
        self.max_unique_words = count_unique_words_with_registry(value + ' ' + main_row, registry)


def get_data_from_workbook(workbook):
    data = []
    for sheet in workbook:
        for row in sheet.iter_rows():
            if row[0].value is not None and type(row[0].value) is str and row[0].value.lower() != 'фраза':
                data.append(row[0].value)

    return data


def clusterize_data(data: list, main_row):
    result = []

    for index, row in enumerate(data):
        if type(row) is str:
            word_registry = create_registry([main_row, row])
            result.append(FilteredRow(row, main_row, word_registry))
        else:
            word_registry = create_registry([main_row, row.value])
            result.append(FilteredRow(row.value, main_row, word_registry))

    return sorted(
        result,
        key=lambda x: (x.max_unique_words, Reverser(len(tokenize(x.value + ' ' + x.main_row)))),
        reverse=True
    )


def write_filtered_row(workbook, data):
    if 'result' in workbook.sheetnames:
        workbook.remove(workbook['result'])

    sheet = workbook.create_sheet('result', 0)

    sheet[f'A1'].value = 'Фраза'
    sheet[f'B1'].value = 'Уникальных слов'
    sheet[f'C1'].value = 'Всего слов'
    sheet[f'D1'].value = 'Дубликатов'
    sheet[f'E1'].value = '% качественных слов'
    for index, row in enumerate(data):
        sheet[f'A{str(index + 2)}'].value = row.value
        sheet[f'B{str(index + 2)}'].value = row.max_unique_words
        sheet[f'C{str(index + 2)}'].value = len(tokenize(row.value+' '+row.main_row))
        sheet[f'D{str(index + 2)}'].value = f'=$C{index + 2}-$B{index + 2}'
        sheet[f'E{str(index + 2)}'].value = f'=ROUND($B{index + 2}/$C{index + 2} * 100)'


def add_key_sheet(workbook, keys_chunks: dict):
    if 'key' in workbook.sheetnames:
        sheet = workbook['key']
    else:
        sheet = workbook.create_sheet('key')
        sheet['A1'].value = 'Фраза'
        sheet['B1'].value = 'Уникальных слов'
        sheet['C1'].value = 'Всего слов'
        sheet['D1'].value = 'Кол-во фраз'
        sheet['E1'].value = '% качество'

    row = 2

    for key_row in keys_chunks.keys():
        sheet[f'B{row}'].value = key_row.max_unique_words
        sheet[f'C{row}'].value = len(tokenize(key_row.value))
        sheet[f'D{row}'].value = len(keys_chunks[key_row])
        sheet[f'E{row}'].value = f'=ROUND($B{row}/$C{row} * 100)'

        for key in keys_chunks[key_row]:
            sheet[f'A{str(row)}'].value = key
            row += 1

        row += 1


def filter_orders_file():
    print(f'{datetime.datetime.now().strftime("%H:%M:%S")} - start')

    workbook = load_workbook(Config.SOURCE_FILE_NAME)
    new_workbook = Workbook()
    new_workbook.remove(new_workbook.active)

    main_row = ''
    keys = []
    keys_chunks = {}

    iteration = 0

    data = clusterize_data(get_data_from_workbook(workbook=workbook), main_row)

    while data:
        print(f'{datetime.datetime.now().strftime("%H:%M:%S")} - iteration {iteration+1}/'
              f'{Config.LIMIT_ITERATIONS or len(data)}, {len(data)} elements left')
        main_row = (main_row + ' ' + data[0].value).strip()
        keys.append(data[0].value)

        data = clusterize_data(data=data[1:], main_row=main_row)

        iteration += 1

        if len(keys) == Config.KEYS_CHUNK_SIZE:
            keys_chunks[FilteredRow(main_row, '', create_registry([main_row]))] = keys
            keys = []
            main_row = ''

        if Config.LIMIT_ITERATIONS and iteration >= Config.LIMIT_ITERATIONS:
            break

    if keys and main_row:
        keys_chunks[FilteredRow(main_row, '', create_registry([main_row]))] = keys

    if data:
        write_filtered_row(new_workbook, data)
        add_key_sheet(
            workbook=new_workbook,
            keys_chunks=keys_chunks,
        )
        new_workbook.save(Config.RESULT_FILE_NAME)

    print(f'{datetime.datetime.now().strftime("%H:%M:%S")} - finished')

